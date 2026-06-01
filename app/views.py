from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from .models import UploadedFile, CleanedDataset, Team, TeamMembership
from .utils.file_parser import parse_file
from .utils.data_cleaner import clean_dataframe
from .utils.auto_chart_suggester import get_chart_suggestions
import json
import os
import math
import pandas as pd
import numpy as np

def register(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            return redirect('dashboard')
    else:
        form = UserCreationForm()
    return render(request, 'registration/register.html', {'form': form})

@login_required
def dashboard(request):
    user_team_ids = TeamMembership.objects.filter(
        user=request.user).values_list('team_id', flat=True)
    files = UploadedFile.objects.filter(
        Q(user=request.user) | 
        Q(team_id__in=user_team_ids)
    ).select_related('cleaneddataset').order_by('-uploaded_at')
    for f in files:
        try:
            f.quality_score = f.cleaneddataset.quality_score
        except Exception:
            f.quality_score = None
    return render(request, 'dashboard.html', {'files': files})

@login_required
def upload_file(request):
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'No file uploaded.')
            return redirect('upload_file')
            
        file = request.FILES['file']
        
        # --- VALIDATION START ---
        ALLOWED_EXTENSIONS = ['csv', 'xlsx', 'xls', 'xlsm', 'json']
        MAX_FILE_SIZE_MB = 50
        MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

        file_ext = file.name.split('.')[-1].lower()
        if file_ext not in ALLOWED_EXTENSIONS:
            messages.error(request, 
                f'File type .{file_ext} is not allowed. '
                f'Please upload: {", ".join(ALLOWED_EXTENSIONS)}')
            return redirect('upload_file')

        if file.size > MAX_FILE_SIZE_BYTES:
            messages.error(request, 
                f'File size {round(file.size/1024/1024, 1)}MB '
                f'exceeds the {MAX_FILE_SIZE_MB}MB limit.')
            return redirect('upload_file')
            
        # Removed MIME type validation to prevent blocking valid Excel files on Windows
        # --- VALIDATION END ---
        
        ext = f".{file_ext}"
            
        custom_name = request.POST.get('custom_name', '')
        
        # Create UploadedFile
        uploaded_file = UploadedFile.objects.create(
            user=request.user,
            original_filename=file.name,
            file_type=ext[1:],
            custom_name=custom_name,
            file=file
        )
        
        # Process file safely using our task
        try:
            from django.conf import settings
            from app.tasks import process_uploaded_file_task
            
            if getattr(settings, 'USE_CELERY', False):
                process_uploaded_file_task.delay(uploaded_file.id)
                messages.success(request, 'File uploaded and is being processed in the background!')
                return redirect('dashboard')
            else:
                result = process_uploaded_file_task(uploaded_file.id)
                if result.get('status') == 'success':
                    messages.success(request, 'File uploaded and cleaned successfully!')
                    return redirect('cleaning_lab', file_id=uploaded_file.id)
                else:
                    raise Exception(result.get('error', 'Unknown error'))
            
        except Exception as e:
            uploaded_file.delete()
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('upload_file')
            
    return render(request, 'upload.html')

def _get_dataset_context(request, file_id):
    user_team_ids = TeamMembership.objects.filter(
        user=request.user).values_list('team_id', flat=True)
    uploaded_file = get_object_or_404(
        UploadedFile.objects.filter(
            Q(user=request.user) | Q(team_id__in=user_team_ids)
        ),
        id=file_id
    )
    dataset = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)
    
    # We NO LONGER inject rows_json into the template to prevent server lag.
    # The frontend will fetch the rows asynchronously via /dataset/<id>/rows/.
    
    return {
        'file_id': file_id,
        'file': uploaded_file,
        'dataset': dataset,
        'columns_json': json.dumps(dataset.columns),
        'unique_columns_json': '[]',
        'stats_json': json.dumps(dataset.stats),
        'raw_snapshot_json': json.dumps(dataset.raw_snapshot),
        'raw_columns_json': json.dumps(dataset.raw_columns),
        'health_report_json': json.dumps(dataset.health_report),
        'quality_score': dataset.quality_score,
        'cleaning_actions_json': json.dumps(dataset.cleaning_actions),
        'before_after_json': json.dumps(dataset.before_after),
        'ai_insights_json': json.dumps(dataset.ai_insights),
        'cell_annotations_json': json.dumps(dataset.cell_annotations),
        'cleaning_options_json': json.dumps(dataset.cleaning_options),
    }

@login_required
def cleaning_lab(request, file_id):
    context = _get_dataset_context(request, file_id)
    return render(request, 'cleaning_lab.html', context)

@login_required
def data_explorer(request, file_id):
    context = _get_dataset_context(request, file_id)
    return render(request, 'data_explorer.html', context)

@login_required
def analytics_studio(request, file_id):
    context = _get_dataset_context(request, file_id)
    return render(request, 'analytics_studio.html', context)

@login_required
def ai_insights_page(request, file_id):
    context = _get_dataset_context(request, file_id)
    return render(request, 'ai_insights_page.html', context)

@login_required
def exports_page(request, file_id):
    context = _get_dataset_context(request, file_id)
    return render(request, 'exports_reports.html', context)


@login_required
def delete_dataset(request, file_id):
    if request.method == 'POST':
        uploaded_file = get_object_or_404(UploadedFile, id=file_id, user=request.user)
        try:
            if uploaded_file.file and os.path.exists(uploaded_file.file.path):
                os.remove(uploaded_file.file.path)
        except Exception:
            pass
        uploaded_file.delete()
        messages.success(request, 'Dataset deleted successfully.')
    return redirect('dashboard')

@login_required
def reclean_dataset(request, file_id):
    """Re-run cleaning pipeline with user-selected toggles."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    uploaded_file = get_object_or_404(UploadedFile, id=file_id, user=request.user)
    
    try:
        options = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON body'}, status=400)
    
    try:
        from django.conf import settings
        from app.tasks import process_uploaded_file_task
        
        if getattr(settings, 'USE_CELERY', False):
            process_uploaded_file_task.delay(uploaded_file.id, options)
            return JsonResponse({'success': True, 'redirect': f'/dataset/{file_id}/'})
        else:
            result = process_uploaded_file_task(uploaded_file.id, options)
            if result.get('status') == 'success':
                return JsonResponse({'success': True, 'redirect': f'/dataset/{file_id}/'})
            else:
                raise Exception(result.get('error', 'Unknown error'))
                
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def dataset_rows_api(request, file_id):
    import traceback
    from django.http import HttpResponse
    from app.utils.parquet_helpers import get_dataframe
    try:
        user_team_ids = TeamMembership.objects.filter(user=request.user).values_list('team_id', flat=True)
        uploaded_file = UploadedFile.objects.filter(
            Q(user=request.user) | Q(team_id__in=user_team_ids)
        ).distinct().get(id=file_id)
    except Exception as e:
        print('FILE ERROR', e)
        raise Http404()
    
    try:
        dataset = CleanedDataset.objects.get(uploaded_file=uploaded_file)
    except Exception as e:
        print('DATASET ERROR', e)
        raise Http404()
    
    try:
        df = get_dataframe(dataset)
        json_str = df.to_json(orient='records', date_format='iso') if not df.empty else '[]'
        return HttpResponse(json_str, content_type='application/json')
    except Exception as e:
        print('DF ERROR', e)
        traceback.print_exc()
        raise Http404()

@login_required
def chart_data(request, file_id):
    """
    JSON API — returns aggregated chart data for a given dataset.

    Query parameters
    ----------------
    x   : str   — X-axis column name (the grouping column)
    y   : str   — Y-axis column name (the value / label column)
    agg : str   — Aggregation mode: sum | count | mean | min | max
                  Default: sum

    Response (JSON array)
    ----------------------
    Sum / Mean / Min / Max mode:
        [{"x_col": "value", "result": 42.0}, ...]

    Count mode:
        [{"x_col": 81, "count": 3, "members": "Ali, Ahmed, Sara"}, ...]

    Why this matters
    ----------------
    The previous frontend code always SUM-aggregated Y values.
    When X = marks and Y = student_name, all students with the same
    mark had their mark values SUMMED (81+81+81 = 243) instead of
    COUNTED (3 students).  This endpoint allows the frontend to request
    the correct aggregation per use-case.
    """
    uploaded_file = get_object_or_404(UploadedFile, id=file_id, user=request.user)
    dataset       = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)

    x_col   = request.GET.get('x', '').strip()
    y_col   = request.GET.get('y', '').strip()
    agg_mode = request.GET.get('agg', 'sum').strip().lower()

    # ── Validate inputs ────────────────────────────────────────────────────
    valid_cols = dataset.columns
    if not x_col or x_col not in valid_cols:
        return JsonResponse({'error': f"Invalid or missing 'x' column: '{x_col}'"}, status=400)
    # Allow '__count__' as a special Y value meaning "count rows"
    if y_col == '__count__':
        agg_mode = 'count'
    elif not y_col or y_col not in valid_cols:
        return JsonResponse(
            {'error': f"Invalid or missing 'y' column: '{y_col}'"}, 
            status=400
        )
    if agg_mode not in ('sum', 'count', 'mean', 'min', 'max'):
        return JsonResponse({'error': f"Invalid agg mode: '{agg_mode}'. Use: sum|count|mean|min|max"}, status=400)

    # ── Rebuild DataFrame from stored rows ─────────────────────────────────
    from app.utils.parquet_helpers import get_dataframe
    df = get_dataframe(dataset)
    if df.empty:
        return JsonResponse([], safe=False)

    # ── Apply aggregation ──────────────────────────────────────────────────
    try:
        if agg_mode == 'count' or y_col == '__count__':
            # For __count__ mode, pass x_col as both args
            # _agg_count will count rows grouped by x_col
            result = _agg_count(df, x_col, x_col)
        else:
            result = _agg_numeric(df, x_col, y_col, agg_mode)
    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=500)

    return JsonResponse(result, safe=False)


# ──────────────────────────────────────────────────────────────────────────────
# AGGREGATION HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _agg_count(df: pd.DataFrame, x_col: str, y_col: str) -> list:
    """
    Group rows by *x_col*, count occurrences, and intelligently collect 
    labels (e.g., student names) for the tooltip.
    """
    label_col = y_col

    # If the user grouped by the same column (e.g. x_col=marks, y_col=marks),
    # or if we just want to be smart about picking the "Name" column:
    if x_col == y_col or pd.api.types.is_numeric_dtype(df[y_col]):
        candidates = ['student_name', 'student', 'name', 'student_id', 'roll_no', 'rollnumber', 'id', 'email']
        # Try to find a matching column (case-insensitive)
        for cand in candidates:
            matches = [c for c in df.columns if c.lower() == cand and c != x_col]
            if matches:
                label_col = matches[0]
                break
        else:
            # Fallback: pick the first non-numeric column that isn't x_col
            non_numeric = df.select_dtypes(exclude='number').columns
            if len(non_numeric) > 0 and non_numeric[0] != x_col:
                label_col = non_numeric[0]
            else:
                # Absolute fallback: pick any column that isn't x_col
                for col in df.columns:
                    if col != x_col:
                        label_col = col
                        break

    grouped = (
        df.groupby(x_col, dropna=False)[label_col]
        .agg(
            count='count',
            members=lambda s: s.dropna().astype(str).tolist()
        )
        .reset_index()
    )

    return [
        {
            x_col:      row[x_col],
            'count':    int(row['count']),
            'students': row['members'],  # Always return under 'students' for the frontend
            'label_col': label_col       # Tell the frontend what column we used
        }
        for _, row in grouped.iterrows()
    ]


def _agg_numeric(df: pd.DataFrame, x_col: str, y_col: str, mode: str) -> list:
    """
    Group rows by *x_col* and apply a numeric aggregation (sum/mean/min/max)
    on *y_col*.

    Coerces *y_col* to numeric first.  Rows where *y_col* cannot be
    converted are dropped.
    """
    df = df.copy()
    df[y_col] = pd.to_numeric(df[y_col], errors='coerce')
    df = df.dropna(subset=[y_col])

    if df.empty:
        raise ValueError(
            f"Column '{y_col}' has no numeric values — "
            f"cannot apply '{mode}' aggregation."
        )

    agg_fn_map = {'sum': 'sum', 'mean': 'mean', 'min': 'min', 'max': 'max'}
    grouped = df.groupby(x_col, dropna=False)[y_col].agg(agg_fn_map[mode]).reset_index()
    grouped.columns = [x_col, 'result']

    return [
        {
            x_col:    row[x_col],
            'result': round(float(row['result']), 4),
        }
        for _, row in grouped.iterrows()
    ]


# ──────────────────────────────────────────────────────────────────────────────
# ADVANCED STATS ENDPOINT
# ──────────────────────────────────────────────────────────────────────────────

@login_required
def advanced_stats(request, file_id):
    """
    JSON API — returns advanced statistical data for charting.

    Query parameters
    ----------------
    col    : str   — Column name to analyse (required)
    type   : str   — Analysis type:
                     histogram | boxplot | violin | correlation
    bins   : int   — Number of histogram bins (default 20, max 100)
    group  : str   — Optional grouping column (for grouped boxplot/violin)
    cols   : str   — Comma-separated list for correlation matrix

    Response shapes
    ---------------
    histogram:
        { bins: [{x0, x1, count, density}], stats: {mean, std, min, max} }

    boxplot:
        { groups: [{name, q1, q2, q3, min, max, mean, outliers:[]}] }

    violin:
        { groups: [{name, kde: [{x, density}], q1, q2, q3, min, max}] }

    correlation:
        { columns: [...], matrix: [[...]] }
    """
    uploaded_file = get_object_or_404(UploadedFile, id=file_id, user=request.user)
    dataset = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)

    col = request.GET.get('col', '').strip()
    analysis_type = request.GET.get('type', 'histogram').strip().lower()
    bins_param = min(int(request.GET.get('bins', 20)), 100)
    group_col = request.GET.get('group', '').strip()
    cols_param = request.GET.get('cols', '').strip()

    valid_cols = dataset.columns

    # Rebuild DataFrame
    from app.utils.parquet_helpers import get_dataframe
    df = get_dataframe(dataset)
    if df.empty:
        return JsonResponse({'error': 'Dataset is empty'}, status=400)

    try:
        if analysis_type == 'histogram':
            if not col or col not in valid_cols:
                return JsonResponse({'error': f"Invalid column: '{col}'"}, status=400)
            result = _compute_histogram(df, col, bins_param)

        elif analysis_type == 'boxplot':
            if not col or col not in valid_cols:
                return JsonResponse({'error': f"Invalid column: '{col}'"}, status=400)
            grp = group_col if group_col and group_col in valid_cols else None
            result = _compute_boxplot(df, col, grp)

        elif analysis_type == 'violin':
            if not col or col not in valid_cols:
                return JsonResponse({'error': f"Invalid column: '{col}'"}, status=400)
            grp = group_col if group_col and group_col in valid_cols else None
            result = _compute_violin(df, col, grp)

        elif analysis_type == 'correlation':
            if cols_param:
                corr_cols = [c.strip() for c in cols_param.split(',') if c.strip() in valid_cols]
            else:
                corr_cols = [c for c in valid_cols if c in df.columns and pd.api.types.is_numeric_dtype(df[c])]
            if len(corr_cols) < 2:
                return JsonResponse({'error': 'Need at least 2 numeric columns for correlation'}, status=400)
            result = _compute_correlation(df, corr_cols)

        else:
            return JsonResponse({'error': f"Unknown analysis type: '{analysis_type}'"}, status=400)

    except Exception as exc:
        return JsonResponse({'error': str(exc)}, status=500)

    return JsonResponse(result, safe=False)


def _safe_float(v):
    """Convert numpy/pandas scalar to Python float, handling NaN/Inf."""
    try:
        f = float(v)
        if math.isnan(f) or math.isinf(f):
            return None
        return round(f, 6)
    except (TypeError, ValueError):
        return None


def _compute_histogram(df: pd.DataFrame, col: str, bins: int) -> dict:
    """Compute histogram bins with count and density."""
    series = pd.to_numeric(df[col], errors='coerce').dropna()
    if series.empty:
        raise ValueError(f"Column '{col}' has no numeric values")

    counts, edges = np.histogram(series.values, bins=bins)
    total = len(series)
    result_bins = []
    for i in range(len(counts)):
        width = edges[i + 1] - edges[i]
        result_bins.append({
            'x0': _safe_float(edges[i]),
            'x1': _safe_float(edges[i + 1]),
            'count': int(counts[i]),
            'density': _safe_float(counts[i] / (total * width)) if total > 0 and width > 0 else 0,
        })

    return {
        'bins': result_bins,
        'stats': {
            'mean': _safe_float(series.mean()),
            'std': _safe_float(series.std()),
            'min': _safe_float(series.min()),
            'max': _safe_float(series.max()),
            'count': total,
        }
    }


def _compute_boxplot(df: pd.DataFrame, col: str, group_col: str | None) -> dict:
    """Compute box plot quartiles, whiskers, and outliers (per group)."""
    df = df.copy()
    df[col] = pd.to_numeric(df[col], errors='coerce')
    df = df.dropna(subset=[col])
    if df.empty:
        raise ValueError(f"Column '{col}' has no numeric values")

    def _box_stats(series, name):
        vals = np.sort(series.values)
        q1 = float(np.percentile(vals, 25))
        q2 = float(np.percentile(vals, 50))
        q3 = float(np.percentile(vals, 75))
        iqr = q3 - q1
        lower_fence = q1 - 1.5 * iqr
        upper_fence = q3 + 1.5 * iqr
        whisker_low = float(vals[vals >= lower_fence].min()) if any(vals >= lower_fence) else q1
        whisker_high = float(vals[vals <= upper_fence].max()) if any(vals <= upper_fence) else q3
        outliers = [_safe_float(v) for v in vals if v < lower_fence or v > upper_fence]
        return {
            'name': str(name),
            'q1': _safe_float(q1), 'q2': _safe_float(q2), 'q3': _safe_float(q3),
            'min': _safe_float(whisker_low), 'max': _safe_float(whisker_high),
            'mean': _safe_float(series.mean()),
            'outliers': outliers[:50],  # cap at 50 outliers for transfer size
        }

    if group_col:
        groups = []
        for name, grp_df in df.groupby(group_col, dropna=False):
            if len(grp_df) > 0:
                groups.append(_box_stats(grp_df[col], name))
        return {'groups': groups}
    else:
        return {'groups': [_box_stats(df[col], col)]}


def _compute_violin(df: pd.DataFrame, col: str, group_col: str | None) -> dict:
    """Compute KDE density for violin plots (per group)."""
    df = df.copy()
    df[col] = pd.to_numeric(df[col], errors='coerce')
    df = df.dropna(subset=[col])
    if df.empty:
        raise ValueError(f"Column '{col}' has no numeric values")

    def _kde_points(series, name):
        vals = series.values
        if len(vals) < 2:
            return {'name': str(name), 'kde': [], 'q1': None, 'q2': None, 'q3': None, 'min': None, 'max': None}

        # Scott's bandwidth
        std = np.std(vals)
        h = 1.06 * std * (len(vals) ** -0.2) if std > 0 else 1.0
        vmin, vmax = vals.min(), vals.max()
        x_pts = np.linspace(vmin - 3 * h, vmax + 3 * h, 80)
        # Gaussian KDE
        kde = np.array([np.mean(np.exp(-0.5 * ((x - vals) / h) ** 2) / (h * np.sqrt(2 * np.pi))) for x in x_pts])

        return {
            'name': str(name),
            'kde': [{'x': _safe_float(x), 'density': _safe_float(d)} for x, d in zip(x_pts, kde)],
            'q1': _safe_float(np.percentile(vals, 25)),
            'q2': _safe_float(np.percentile(vals, 50)),
            'q3': _safe_float(np.percentile(vals, 75)),
            'min': _safe_float(vals.min()),
            'max': _safe_float(vals.max()),
        }

    if group_col:
        groups = []
        for name, grp_df in df.groupby(group_col, dropna=False):
            if len(grp_df) > 0:
                groups.append(_kde_points(grp_df[col], name))
        return {'groups': groups}
    else:
        return {'groups': [_kde_points(df[col], col)]}


def _compute_correlation(df: pd.DataFrame, cols: list) -> dict:
    """Compute Pearson correlation matrix for given columns."""
    sub = df[cols].apply(pd.to_numeric, errors='coerce').dropna()
    if sub.shape[0] < 2:
        raise ValueError("Not enough rows with numeric values to compute correlation")
    corr = sub.corr(method='pearson')
    matrix = [[_safe_float(corr.iloc[i][j]) for j in range(len(cols))] for i in range(len(cols))]
    return {'columns': cols, 'matrix': matrix}

def get_unique_columns(rows):
    if not rows:
        return []
    import pandas as pd
    df = pd.DataFrame(rows)
    unique_columns = []
    for column in df.columns:
        if df[column].nunique() == len(df):
            unique_columns.append(column)
    return unique_columns

@login_required
def data_profiler(request, file_id):
    uploaded_file = get_object_or_404(UploadedFile, id=file_id, user=request.user)
    dataset = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)
    from app.utils.parquet_helpers import get_dataframe
    df = get_dataframe(dataset)
    unique_cols = get_unique_columns(df.to_dict(orient='records') if not df.empty else [])
    context = {
        'file': uploaded_file,
        'dataset': dataset,
        'file_id': file_id,
        'columns_json': json.dumps(dataset.columns),
        'rows_json': df.to_json(orient='records', date_format='iso') if not df.empty else '[]',
        'unique_columns_json': json.dumps(unique_cols),
        'stats_json': json.dumps(dataset.stats),
        'health_report_json': json.dumps(dataset.health_report),
        'quality_score': dataset.quality_score,
    }
    return render(request, 'data_profiler.html', context)

@login_required
def chart_suggestions(request, file_id):
    """API endpoint to get auto chart suggestions based on dataset schema."""
    uploaded_file = get_object_or_404(UploadedFile, id=file_id, user=request.user)
    dataset = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)
    
    from app.utils.auto_chart_suggester import get_chart_suggestions
    suggestions = get_chart_suggestions(dataset)
    
    return JsonResponse({'suggestions': suggestions})

from django.http import HttpResponse
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.template.loader import render_to_string

@login_required
def export_excel(request, file_id):
    uploaded_file = get_object_or_404(UploadedFile, id=file_id, user=request.user)
    dataset = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)

    wb = openpyxl.Workbook()
    
    # Sheet 1: Cleaned Data
    ws1 = wb.active
    ws1.title = "Cleaned Data"
    
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    ws1.append(dataset.columns)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    from app.utils.parquet_helpers import get_dataframe
    df = get_dataframe(dataset)
    rows_data = df.to_dict(orient='records') if not df.empty else []
    for row in rows_data:
        ws1.append([row.get(col, "") for col in dataset.columns])
        
    # Sheet 2: Stats
    ws2 = wb.create_sheet(title="Stats Summary")
    ws2.append(["Column", "Count", "Unique", "Top", "Mean", "Min", "Max"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for col, stat in dataset.stats.items():
        ws2.append([
            col,
            stat.get('count', ''),
            stat.get('unique', ''),
            stat.get('top', ''),
            stat.get('mean', ''),
            stat.get('min', ''),
            stat.get('max', '')
        ])

    # Sheet 3: Cleaning Log
    ws3 = wb.create_sheet(title="Cleaning Log")
    ws3.append(["Action"])
    ws3["A1"].fill = header_fill
    ws3["A1"].font = header_font
    for log in dataset.cleaning_log:
        ws3.append([log])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="data_pro_{file_id}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_pdf(request, file_id):
    import weasyprint
    uploaded_file = get_object_or_404(UploadedFile, id=file_id, user=request.user)
    dataset = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)
    
    context = {
        'file': uploaded_file,
        'quality_score': dataset.quality_score,
        'columns': dataset.columns,
        'rows_preview': get_dataframe(dataset).to_dict(orient='records')[:20] if not get_dataframe(dataset).empty else [],
        'stats': dataset.stats,
        'cleaning_log': dataset.cleaning_log,
    }
    html_string = render_to_string('pdf_report.html', context)
    
    pdf_file = weasyprint.HTML(string=html_string).write_pdf()
    
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="data_pro_report_{file_id}.pdf"'
    return response

from app.utils.cleaning_engine import DataCleaningEngine

@login_required
def cleaning_center_view(request, file_id):
    from app.utils.parquet_helpers import get_dataframe
    context = _get_dataset_context(request, file_id)
    
    # Inject rows and unique columns specifically for the SPA
    dataset = context['dataset']
    df = get_dataframe(dataset)
    
    # Calculate unique values for categorical columns (useful for Analytics and Filtering)
    unique_columns = []
    if not df.empty:
        for col in df.columns:
            if df[col].nunique() < 50:
                unique_columns.append({
                    "column": col,
                    "values": [x for x in df[col].dropna().unique().tolist() if x != ""]
                })
    
    context['rows_json'] = json.dumps(df.to_dict(orient='records'), default=str)
    context['unique_columns_json'] = json.dumps(unique_columns, default=str)
    
    return render(request, 'cleaning_center.html', context)

@login_required
def api_preview_cleaning(request, file_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'error': 'Only POST allowed'}, status=405)
        
    dataset = get_object_or_404(CleanedDataset, uploaded_file__id=file_id)
    
    try:
        payload = json.loads(request.body)
        pipeline = payload.get('pipeline', [])
        
        df_raw = pd.DataFrame(dataset.raw_snapshot)
        if df_raw.empty:
            return JsonResponse({'status': 'error', 'error': 'No raw snapshot available for preview.'}, status=400)
            
        engine = DataCleaningEngine()
        df_cleaned, logs = engine.apply_pipeline(df_raw, pipeline)
        
        # Replace NaN/inf to avoid JSON serialization errors
        df_raw.replace([np.inf, -np.inf], np.nan, inplace=True)
        df_cleaned.replace([np.inf, -np.inf], np.nan, inplace=True)
        
        before_records = df_raw.head(100).fillna("").to_dict(orient='records')
        after_records = df_cleaned.head(100).fillna("").to_dict(orient='records')
        
        return JsonResponse({
            'status': 'success',
            'before': before_records,
            'after': after_records,
            'logs': logs,
            'before_columns': list(df_raw.columns),
            'after_columns': list(df_cleaned.columns),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'error': str(e)}, status=400)

@login_required
def api_apply_cleaning(request, file_id):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'error': 'Only POST allowed'}, status=405)
        
    from app.utils.parquet_helpers import get_dataframe, save_dataframe_to_parquet
    
    dataset = get_object_or_404(CleanedDataset, uploaded_file__id=file_id)
    
    try:
        payload = json.loads(request.body)
        pipeline = payload.get('pipeline', [])
        
        df_full = get_dataframe(dataset)
        if df_full.empty:
            return JsonResponse({'status': 'error', 'error': 'Dataset is empty.'}, status=400)
            
        engine = DataCleaningEngine()
        df_cleaned, logs = engine.apply_pipeline(df_full, pipeline)
        
        dataset.columns = list(df_cleaned.columns)
        
        history = list(dataset.pipeline_history) if dataset.pipeline_history else []
        history.append({
            'pipeline': pipeline,
            'logs': logs
        })
        dataset.pipeline_history = history
        
        # Overwrite current logs with new execution logs
        dataset.cleaning_log = logs
        dataset.save()
        
        save_dataframe_to_parquet(df_cleaned, dataset)
        
        dataset.uploaded_file.row_count = len(df_cleaned)
        dataset.uploaded_file.column_count = len(df_cleaned.columns)
        dataset.uploaded_file.save()
        
        return JsonResponse({'status': 'success', 'message': 'Cleaning applied successfully.', 'logs': logs})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'status': 'error', 'error': str(e)}, status=400)


@login_required
def chart_suggestions(request, file_id):
    """Return AI-generated chart suggestions for the given dataset as JSON."""
    if request.method != 'GET':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    uploaded_file = get_object_or_404(
        UploadedFile.objects.filter(
            Q(user=request.user) | Q(team__teammembership__user=request.user)
        ).distinct(),
        id=file_id,
    )
    dataset = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)

    try:
        suggestions = get_chart_suggestions(dataset)
        return JsonResponse({'suggestions': suggestions, 'count': len(suggestions)})
    except Exception as e:
        return JsonResponse({'error': str(e), 'suggestions': []}, status=500)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def export_excel(request, file_id):
    """Generate a multi-sheet .xlsx and return it as a file download."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from app.utils.parquet_helpers import get_dataframe

    uploaded_file = get_object_or_404(
        UploadedFile.objects.filter(
            Q(user=request.user) | Q(team__teammembership__user=request.user)
        ).distinct(),
        id=file_id,
    )
    dataset = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)

    wb = Workbook()

    # ── Sheet 1: Cleaned Data ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Cleaned Data'

    df = get_dataframe(dataset)
    columns = list(df.columns) if not df.empty else (dataset.columns or [])

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill('solid', fgColor='1E6F50')   # dark emerald
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=str(col_name))
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        ws1.column_dimensions[get_column_letter(col_idx)].width = max(len(str(col_name)) + 4, 12)

    if not df.empty:
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                ws1.cell(row=row_idx, column=col_idx, value=value)
    ws1.freeze_panes = 'A2'   # keep header visible when scrolling

    # ── Sheet 2: Stats ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Stats')
    stats = dataset.stats or {}

    stat_keys = ['count', 'mean', 'std', 'min', '25%', '50%', '75%', 'max',
                 'unique', 'top', 'freq', 'sum', 'median', 'variance']

    # Header row
    header_row = ['Column'] + stat_keys
    for col_idx, h in enumerate(header_row, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font  = Font(bold=True, color='FFFFFF')
        cell.fill  = PatternFill('solid', fgColor='2563EB')   # blue
        cell.alignment = Alignment(horizontal='center')
        ws2.column_dimensions[get_column_letter(col_idx)].width = 14

    for row_idx, (col_name, col_stats) in enumerate(stats.items(), start=2):
        ws2.cell(row=row_idx, column=1, value=col_name)
        for col_idx, key in enumerate(stat_keys, start=2):
            ws2.cell(row=row_idx, column=col_idx, value=col_stats.get(key, ''))

    # ── Sheet 3: Cleaning Log ──────────────────────────────────────────────
    ws3 = wb.create_sheet('Cleaning Log')
    log_header = ws3.cell(row=1, column=1, value='Cleaning Actions')
    log_header.font  = Font(bold=True, color='FFFFFF')
    log_header.fill  = PatternFill('solid', fgColor='7C3AED')   # violet
    ws3.column_dimensions['A'].width = 80

    cleaning_log = dataset.cleaning_log or []
    for row_idx, entry in enumerate(cleaning_log, start=2):
        ws3.cell(row=row_idx, column=1, value=str(entry))

    # ── Serialise to memory and return ────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = uploaded_file.original_filename.rsplit('.', 1)[0]
    filename  = f'{safe_name}_cleaned.xlsx'

    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# PDF EXPORT
# ─────────────────────────────────────────────────────────────────────────────

@login_required
def export_pdf(request, file_id):
    """Render a PDF report with reportlab and return it as a file download."""
    import io
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from app.utils.parquet_helpers import get_dataframe

    uploaded_file = get_object_or_404(
        UploadedFile.objects.filter(
            Q(user=request.user) | Q(team__teammembership__user=request.user)
        ).distinct(),
        id=file_id,
    )
    dataset = get_object_or_404(CleanedDataset, uploaded_file=uploaded_file)

    df           = get_dataframe(dataset)
    columns      = list(df.columns) if not df.empty else (dataset.columns or [])
    rows_preview = df.head(20).to_dict(orient='records') if not df.empty else []
    stats        = dataset.stats or {}
    cleaning_log = dataset.cleaning_log or []
    quality      = dataset.quality_score or 0
    dataset_name = uploaded_file.custom_name or uploaded_file.original_filename

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=18)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    elements.append(Paragraph(f"<b>Dataset Report: {dataset_name}</b>", styles['Title']))
    elements.append(Spacer(1, 10))
    
    # Meta Info
    meta_text = f"Rows: {uploaded_file.row_count} | Columns: {uploaded_file.column_count} | Quality Score: {quality}%"
    elements.append(Paragraph(meta_text, styles['Normal']))
    elements.append(Spacer(1, 20))

    # Data Preview
    elements.append(Paragraph("<b>Data Preview (First 20 Rows)</b>", styles['Heading2']))
    elements.append(Spacer(1, 10))

    # Convert preview data into list of lists
    preview_data = [columns]
    for row in rows_preview:
        preview_data.append([str(row.get(col, ""))[:30] for col in columns]) # Trucate long cells
    
    if len(preview_data) > 1:
        t_preview = Table(preview_data, repeatRows=1)
        t_preview.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f5f7ff')),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#dddddd')),
            ('FONTSIZE', (0,1), (-1,-1), 8),
        ]))
        elements.append(t_preview)
    else:
        elements.append(Paragraph("No data available.", styles['Normal']))
    
    elements.append(Spacer(1, 20))

    # Summary Statistics
    elements.append(Paragraph("<b>Summary Statistics</b>", styles['Heading2']))
    elements.append(Spacer(1, 10))
    stat_keys = ['count', 'mean', 'std', 'min', 'max', 'unique', 'top']
    stats_data = [['Column'] + stat_keys]
    for col_name, col_stats in stats.items():
        row = [col_name] + [str(col_stats.get(k, "-"))[:15] for k in stat_keys]
        stats_data.append(row)
    
    if len(stats_data) > 1:
        t_stats = Table(stats_data, repeatRows=1)
        t_stats.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 8),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f5f7ff')),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#dddddd')),
            ('FONTSIZE', (0,1), (-1,-1), 8),
        ]))
        elements.append(t_stats)
    else:
        elements.append(Paragraph("No statistics available.", styles['Normal']))

    elements.append(Spacer(1, 20))

    # Cleaning Log
    elements.append(Paragraph("<b>Cleaning Actions</b>", styles['Heading2']))
    elements.append(Spacer(1, 10))
    if cleaning_log:
        for entry in cleaning_log:
            elements.append(Paragraph(f"• {entry}", styles['Normal']))
    else:
        elements.append(Paragraph("No cleaning actions recorded.", styles['Normal']))

    doc.build(elements)
    pdf_bytes = buffer.getvalue()
    buffer.close()

    safe_name = uploaded_file.original_filename.rsplit('.', 1)[0]
    filename  = f'{safe_name}_report.pdf'

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def create_team(request):
    pass

@login_required
def invite_member(request, team_id):
    pass

@login_required
def team_settings(request, team_id):
    pass

@login_required
def sheet_viewer(request, file_id):
    """Multi-sheet viewer — shows all sheets from an Excel file."""
    uploaded_file = get_object_or_404(
        UploadedFile, id=file_id, user=request.user
    )
    dataset = get_object_or_404(
        CleanedDataset, uploaded_file=uploaded_file
    )

    all_sheets = dataset.all_sheets_data or []
    sheet_idx = int(request.GET.get('sheet', 0))
    if sheet_idx >= len(all_sheets) or sheet_idx < 0:
        sheet_idx = 0

    current_sheet = all_sheets[sheet_idx] if all_sheets else None

    return render(request, 'app/sheet_viewer.html', {
        'uploaded_file': uploaded_file,
        'dataset': dataset,
        'all_sheets': all_sheets,
        'current_sheet': current_sheet,
        'current_sheet_idx': sheet_idx,
        'total_sheets': len(all_sheets),
    })
